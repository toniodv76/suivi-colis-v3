import os
import sqlite3
from datetime import date
from io import BytesIO

import psycopg2
from flask import Flask, request, redirect, render_template, jsonify, send_file, abort
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)

ACCESS_CODE = os.environ.get("ACCESS_CODE", "h2otech")
DATABASE_URL = os.environ.get("DATABASE_URL")
LOCAL_DB = "colis.db"


def is_postgres():
    return bool(DATABASE_URL and DATABASE_URL.startswith(("postgres://", "postgresql://")))


def get_conn():
    if is_postgres():
        return psycopg2.connect(DATABASE_URL)
    return sqlite3.connect(LOCAL_DB)


def placeholder():
    return "%s" if is_postgres() else "?"


def rows_to_dicts(cur):
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]


def init_db():
    conn = get_conn()
    cur = conn.cursor()

    if is_postgres():
        cur.execute("""
            CREATE TABLE IF NOT EXISTS colis (
                id SERIAL PRIMARY KEY,
                client TEXT NOT NULL,
                adresse TEXT NOT NULL,
                contenu TEXT NOT NULL,
                statut TEXT NOT NULL DEFAULT 'A FAIRE',
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                done_at TIMESTAMP NULL
            )
        """)
        # migration si ancienne base
        cur.execute("ALTER TABLE colis ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP")
        cur.execute("ALTER TABLE colis ADD COLUMN IF NOT EXISTS done_at TIMESTAMP NULL")
    else:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS colis (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                client TEXT NOT NULL,
                adresse TEXT NOT NULL,
                contenu TEXT NOT NULL,
                statut TEXT NOT NULL DEFAULT 'A FAIRE',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                done_at TEXT NULL
            )
        """)
        # migration SQLite simple
        cur.execute("PRAGMA table_info(colis)")
        cols = [r[1] for r in cur.fetchall()]
        if "updated_at" not in cols:
            cur.execute("ALTER TABLE colis ADD COLUMN updated_at TEXT DEFAULT CURRENT_TIMESTAMP")
        if "done_at" not in cols:
            cur.execute("ALTER TABLE colis ADD COLUMN done_at TEXT NULL")

    conn.commit()
    conn.close()


def require_code():
    code = request.values.get("code", "")
    if code != ACCESS_CODE:
        abort(403)


def get_colis(include_done=True, month_start=None, month_end=None):
    conn = get_conn()
    cur = conn.cursor()
    params = []
    where = []

    if not include_done:
        where.append("statut <> 'FAIT'")

    if month_start and month_end:
        ph = placeholder()
        where.append(f"created_at >= {ph}")
        params.append(month_start)
        where.append(f"created_at < {ph}")
        params.append(month_end)

    sql = """
        SELECT id, client, adresse, contenu, statut, created_at, updated_at, done_at
        FROM colis
    """

    if where:
        sql += " WHERE " + " AND ".join(where)

    sql += """
        ORDER BY
        CASE
            WHEN statut='A FAIRE' THEN 0
            WHEN statut='EN COURS' THEN 1
            WHEN statut='FAIT' THEN 2
            ELSE 3
        END,
        created_at DESC
    """

    cur.execute(sql, params)
    data = rows_to_dicts(cur)
    conn.close()
    return data


def month_bounds(target=None):
    target = target or date.today()
    start = date(target.year, target.month, 1)
    if target.month == 12:
        end = date(target.year + 1, 1, 1)
    else:
        end = date(target.year, target.month + 1, 1)
    return start, end


@app.route("/", methods=["GET", "POST"])
def index():
    code = request.values.get("code", "")
    if code != ACCESS_CODE:
        return render_template("login.html")

    if request.method == "POST":
        client = request.form.get("client", "").strip()
        adresse = request.form.get("adresse", "").strip()
        contenu = request.form.get("contenu", "").strip()
        statut = request.form.get("statut", "A FAIRE").strip().upper()

        if statut not in ["A FAIRE", "EN COURS", "FAIT"]:
            statut = "A FAIRE"

        if client and adresse and contenu:
            conn = get_conn()
            cur = conn.cursor()
            ph = placeholder()
            done_value = "CURRENT_TIMESTAMP" if statut == "FAIT" else "NULL"

            cur.execute(
                f"""
                INSERT INTO colis (client, adresse, contenu, statut, updated_at, done_at)
                VALUES ({ph}, {ph}, {ph}, {ph}, CURRENT_TIMESTAMP, {done_value})
                """,
                (client, adresse, contenu, statut),
            )
            conn.commit()
            conn.close()

        return redirect(f"/?code={ACCESS_CODE}")

    colis = get_colis(include_done=True)
    return render_template("index.html", colis=colis, code=ACCESS_CODE)


@app.route("/edit/<int:colis_id>", methods=["GET", "POST"])
def edit(colis_id):
    require_code()
    conn = get_conn()
    cur = conn.cursor()
    ph = placeholder()

    if request.method == "POST":
        client = request.form.get("client", "").strip()
        adresse = request.form.get("adresse", "").strip()
        contenu = request.form.get("contenu", "").strip()
        statut = request.form.get("statut", "A FAIRE").strip().upper()

        if statut not in ["A FAIRE", "EN COURS", "FAIT"]:
            statut = "A FAIRE"

        done_sql = "CURRENT_TIMESTAMP" if statut == "FAIT" else "NULL"

        cur.execute(
            f"""
            UPDATE colis
            SET client={ph}, adresse={ph}, contenu={ph}, statut={ph},
                updated_at=CURRENT_TIMESTAMP, done_at={done_sql}
            WHERE id={ph}
            """,
            (client, adresse, contenu, statut, colis_id),
        )
        conn.commit()
        conn.close()
        return redirect(f"/?code={ACCESS_CODE}")

    cur.execute(f"SELECT id, client, adresse, contenu, statut, created_at, updated_at, done_at FROM colis WHERE id={ph}", (colis_id,))
    rows = rows_to_dicts(cur)
    conn.close()

    if not rows:
        abort(404)

    return render_template("edit.html", item=rows[0], code=ACCESS_CODE)


@app.route("/statut/<int:colis_id>/<statut>", methods=["POST", "GET"])
def statut(colis_id, statut):
    require_code()
    statut = statut.upper().replace("-", " ")

    if statut not in ["A FAIRE", "EN COURS", "FAIT"]:
        abort(400)

    conn = get_conn()
    cur = conn.cursor()
    ph = placeholder()
    done_sql = "CURRENT_TIMESTAMP" if statut == "FAIT" else "NULL"

    cur.execute(
        f"UPDATE colis SET statut={ph}, updated_at=CURRENT_TIMESTAMP, done_at={done_sql} WHERE id={ph}",
        (statut, colis_id),
    )
    conn.commit()
    conn.close()
    return redirect(f"/?code={ACCESS_CODE}")


@app.route("/delete/<int:colis_id>", methods=["POST", "GET"])
def delete(colis_id):
    require_code()
    conn = get_conn()
    cur = conn.cursor()
    ph = placeholder()
    cur.execute(f"DELETE FROM colis WHERE id={ph}", (colis_id,))
    conn.commit()
    conn.close()
    return redirect(f"/?code={ACCESS_CODE}")


@app.route("/historique")
def historique():
    require_code()
    colis = get_colis(include_done=True)
    return render_template("historique.html", colis=colis, code=ACCESS_CODE)


@app.route("/api/colis")
def api_colis():
    return jsonify(get_colis(include_done=True))


@app.route("/ecran")
def ecran():
    colis = get_colis(include_done=True)
    return render_template("ecran.html", colis=colis)


def build_excel(rows, title="Suivi colis"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Envois colis"

    ws.append([title])
    ws.merge_cells("A1:H1")
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append(["Date création", "Client", "Adresse", "Colis / description", "Statut", "Dernière modif", "Date fait", "ID"])

    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    for r in rows:
        ws.append([
            str(r.get("created_at", "")),
            r.get("client", ""),
            r.get("adresse", ""),
            r.get("contenu", ""),
            r.get("statut", ""),
            str(r.get("updated_at", "")),
            str(r.get("done_at", "")),
            r.get("id", ""),
        ])

    widths = [22, 25, 40, 60, 16, 22, 22, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    for row in ws.iter_rows(min_row=3):
        status = row[4].value
        if status == "FAIT":
            color = "C6EFCE"
        elif status == "EN COURS":
            color = "BDD7EE"
        else:
            color = "FFC7CE"

        fill = PatternFill("solid", fgColor=color)

        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = fill

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


@app.route("/export/monthly")
def export_monthly():
    require_code()
    start, end = month_bounds()
    rows = get_colis(include_done=True, month_start=start.isoformat(), month_end=end.isoformat())
    bio = build_excel(rows, f"Envois colis - {start.strftime('%m/%Y')}")
    return send_file(
        bio,
        as_attachment=True,
        download_name=f"envois_colis_{start.strftime('%Y_%m')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/export/all")
def export_all():
    require_code()
    rows = get_colis(include_done=True)
    bio = build_excel(rows, "Historique complet envois colis")
    return send_file(
        bio,
        as_attachment=True,
        download_name="historique_complet_envois_colis.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=False)
